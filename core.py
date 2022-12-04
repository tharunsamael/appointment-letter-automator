from re import template
from docx import Document
import openpyxl
from xlwt import Workbook
import shutil
from docx2pdf import convert
import sys

def rec_func(document, word, replacing_word):
    for paragraph in document.paragraphs:
        if(word in paragraph.text):
            inline = paragraph.runs
            for j in range(len(inline)):
                if (word in inline[j].text):
                    text = inline[j].text.replace(word, replacing_word)
                    inline[j].text = text
    return "completed"


def automtor(template_file, data_file, replacing_words_length, replacing_words_list):
    replacing_words_list = replacing_words_list.split(",")
    
    new_replacing_words_list = []
    for i in replacing_words_list:
        new_replacing_words_list.append(i.strip())
    
    wb = Workbook()
    sheet1 = wb.add_sheet('Sheet 1')

    wrkbk = openpyxl.load_workbook(fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\uploads\{data_file}")
    
    sh = wrkbk.active
    
    filled_rows = len(sh['A'])
    filled_columns = int(replacing_words_length)

    li = []
    for row in sh.iter_rows(min_row=1, min_col=1, max_col=filled_columns, max_row=filled_rows):
        internal_li = []
        for cell in row:
            internal_li.append(cell.value)
        li.append(internal_li)

    for i in li:
        shutil.copyfile(fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\uploads\{template_file}", fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\wordDocs\{i[0]}"+".docx")
        document = Document(fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\wordDocs\{i[0]}"+".docx")
        for func in range(int(replacing_words_length)):
            rec_func(document, new_replacing_words_list[func], i[func])
        document.save(fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\wordDocs\{i[0]}"+".docx")
        convert(fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\wordDocs\{i[0]}"+".docx", fr"C:\Users\Puliyang\Documents\5th Sem\newprojbackend\pdfDocs\{i[0]}"+".pdf")
    return("done")

print(automtor(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]))